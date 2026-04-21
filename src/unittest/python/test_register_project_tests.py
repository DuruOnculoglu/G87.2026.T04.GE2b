"""class for testing the register_order method"""
import unittest
import os
import json
import freezegun

from openpyxl import load_workbook
from uc3m_consulting.enterprise_manager import EnterpriseManager
from uc3m_consulting.enterprise_management_exception import EnterpriseManagementException
from freezegun import freeze_time


def create_temp_json(file_name, content):
    file_path = os.path.join("src/unittest/python", file_name)

    content = json.loads(content)

    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(content, f, indent=4)

    return file_path


class MyTestCase(unittest.TestCase):
    """class for testing the register_order method"""

    @classmethod
    def setUpClass(cls):
        cls.mngr = EnterpriseManager()

        wb = load_workbook("docs/GE2ExcelTestFile.xlsx")
        sheet = wb.active

        cls.headers = [cell.value for cell in sheet[1]]
        cls.rows = list(sheet.iter_rows(min_row=2, values_only=True))

    @freeze_time("2026/01/01")
    def test_cases_from_excel(self):
        for row in self.rows:
            row_dict = dict(zip(self.headers, row))

            test_id = row_dict["ID TEST"]
            test_type = row_dict["TYPE (DUPLICATION / DELETION / MODIFICATION / VALID)"]
            file_path = create_temp_json(
                row_dict["FILE PATH"],
                row_dict["FILE CONTENT"])
            expected = row_dict["EXPECTED RESULT"]

            self.addCleanup(lambda p=file_path: os.remove(p) if os.path.exists(p) else None)

            if test_type == "VALID":
                with self.subTest(test_id=test_id):
                    result = self.mngr.register_document(file_path)
                    self.assertEqual(expected, result)
            else:
                with self.subTest(test_id=test_id):
                    expected_message = row_dict["EXPECTED RESULT"]
                    with self.assertRaises(EnterpriseManagementException) as context:
                        self.mngr.register_document(file_path)
                    actual_message = str(context.exception)
                    self.assertEqual(expected_message, actual_message)

    def tearDown(self):
         pass

if __name__ == '__main__':
    unittest.main()

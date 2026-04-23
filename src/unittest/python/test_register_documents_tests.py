"""class for testing the register_order method"""
import unittest
import os
import json

from openpyxl import load_workbook
from uc3m_consulting.enterprise_manager import EnterpriseManager
from uc3m_consulting.enterprise_management_exception import EnterpriseManagementException
from freezegun import freeze_time


def create_temp_json(file_name, content):
    if not file_name:
        raise ValueError("FILE PATH is missing in Excel test case")

    file_path = os.path.join("src/unittest/python", file_name)

    if not content:
        return file_path

    try:
        parsed = json.loads(content)
    except json.JSONDecodeError:
        parsed = content

    with open(file_path, "w", encoding="utf-8") as f:
        if isinstance(parsed, (dict, list)):
            json.dump(parsed, f, indent=4)
        else:
            f.write(str(parsed))

    return file_path


class MyTestCase(unittest.TestCase):
    """class for testing the register_order method"""

    @classmethod
    def setUpClass(cls):
        cls.mngr = EnterpriseManager()

        wb = load_workbook("docs/GE2ExcelTestFile.xlsx")
        sheet = wb.active

        cls.headers = [cell.value for cell in sheet[1]]
        cls.rows = list(sheet.iter_rows(min_row=2, values_only=True))

    @freeze_time("2026/01/01")
    def test_cases_from_xlsx(self):
        for row in self.rows:
            if all(cell is None for cell in row):
                continue

            row_dict = dict(zip(self.headers, row))

            test_id = row_dict["ID TEST"]
            test_type = row_dict["TYPE"]
            file_name = row_dict.get("FILE PATH")
            if not file_name:
                raise ValueError(f"Test {test_id}: FILE PATH is missing")

            file_path = create_temp_json(
                row_dict["FILE PATH"],
                row_dict["FILE CONTENT"])
            expected = row_dict["EXPECTED RESULT"]

            self.addCleanup(lambda p=file_path: os.remove(p) if os.path.exists(p) else None)

            if test_type == "VALID":
                with self.subTest(test_id=test_id):
                    result = self.mngr.register_document(file_path)
                    self.assertEqual(expected, result)
            else:
                with self.subTest(test_id=test_id):
                    expected_message = row_dict["EXPECTED RESULT"]
                    with self.assertRaises(EnterpriseManagementException) as context:
                        self.mngr.register_document(file_path)
                    actual_message = str(context.exception)
                    self.assertEqual(expected_message, actual_message)

    # def test_case_76(self):
    #     """ This test case tests the internal processing error"""
    #     valid_data = ' {"PROJECT_ID":"00112233445566778899AaBbCcDdEeFf","FILENAME":"File1357.pdf"} '
    #     expected_message = "Internal processing error when getting the file_signature."
    #
    #     test_file = create_temp_json("test76.json", valid_data)
    #
    #     storage_file = "src/main/python/uc3m_consulting/all_documents.json"
    #
    #     if os.path.exists(storage_file):
    #         os.remove(storage_file)
    #
    #     self.mngr.project_id = "INVALID"
    #
    #     try:
    #         with self.assertRaises(EnterpriseManagementException) as context:
    #             self.mngr.register_document(test_file)
    #         self.assertEqual(expected_message, str(context.exception))
    #
    #     finally:
    #         if os.path.exists(storage_file):
    #             os.remove(storage_file)

    def tearDown(self):
         pass

if __name__ == '__main__':
    unittest.main()
